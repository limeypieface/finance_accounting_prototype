"""
XLSX source adapter for journal and list exports (e.g. QuickBooks Excel export).

Supports flexible layout:
  - sheet by index (0-based) or name
  - header row by index or auto-detect (scans first N rows for journal-like column names)
  - skip_rows before header
  - normalizes cell values (strip, blank->empty string)

Auto-detect looks for a row containing at least 2 of: date, account, debit, credit,
amount, description, memo, reference, name, type (so varied journal/export layouts match).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterator

from finance_ingestion.adapters.base import SourceProbe

# Journal-like column keywords (normalized: strip, lower); row with ≥2 matches is header candidate
_HEADER_KEYWORDS = frozenset({
    "date", "transaction date", "trans date", "post date", "journal date",
    "account", "account name", "account number", "account code", "name", "account type",
    "debit", "credits", "credit", "amount", "balance",
    "description", "memo", "reference", "details", "notes",
    "type", "detail type", "full name",
})


def _normalize_header_cell(value: Any) -> str:
    """Normalize a cell value for header matching or key use."""
    if value is None:
        return ""
    s = str(value).strip()
    # Drop common Excel artifacts and collapse spaces
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _cell_value(row: Any, col_idx: int) -> Any:
    """Get cell value from openpyxl row (0-based column index)."""
    try:
        cell = row[col_idx]
        if cell is None:
            return ""
        v = cell.value
        if v is None:
            return ""
        if isinstance(v, float):
            if v == int(v):
                return int(v)
            return v
        return str(v).strip() if v else ""
    except (IndexError, TypeError):
        return ""


def _row_to_keywords(row: Any, max_cols: int = 30) -> set[str]:
    """Extract normalized keywords from a row for header scoring."""
    keywords = set()
    for c in range(max_cols):
        v = _cell_value(row, c)
        if not v:
            continue
        v_lower = v.lower().strip()
        if v_lower in _HEADER_KEYWORDS:
            keywords.add(v_lower)
        # Also check if any keyword is a substring (e.g. "Debit Amount")
        for kw in _HEADER_KEYWORDS:
            if kw in v_lower or v_lower in kw:
                keywords.add(kw)
    return keywords


def _detect_header_row(rows: list, max_search: int = 15, min_keywords: int = 2) -> int:
    """Return 0-based row index of the first row that looks like a header (journal columns)."""
    for i, row in enumerate(rows[:max_search]):
        kw = _row_to_keywords(row)
        if len(kw) >= min_keywords:
            return i
    return 0


def _column_count(row: Any) -> int:
    """Number of non-empty trailing cells we consider part of the row."""
    n = 0
    for c in range(50):
        v = _cell_value(row, c)
        if v != "" and v is not None:
            n = c + 1
    return max(n, 1)


class XlsxSourceAdapter:
    """
    Read .xlsx files as one dict per row. Uses first row (or auto-detected header row)
    as column names.

    source_options:
      sheet: 0-based sheet index (int) or sheet name (str). Default: active sheet.
      skip_rows: number of rows to skip at top of sheet before header/data. Default: 0.
      header_row: 0-based row index (within the sheet after skip_rows) to use as header.
        If omitted and auto_detect_header is true, the first row that contains at least 2
        journal-like column names (date, account, debit, credit, amount, description, ...) is used.
      auto_detect_header: if true (default), scan first 15 rows to find a header row; if false, use header_row or row 0.
    """

    def read(self, source_path: Path, options: dict[str, Any]) -> Iterator[dict[str, Any]]:
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("XLSX support requires openpyxl. Install with: pip install openpyxl") from e

        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet = self._get_sheet(wb, options)
            skip_rows = int(options.get("skip_rows", 0))
            header_row_idx = options.get("header_row")
            auto_detect = options.get("auto_detect_header", True)

            # Build list of rows (read_only sheet iterates once)
            rows = list(sheet.iter_rows(min_row=1 + skip_rows, max_row=100_000))
            if not rows:
                return

            if header_row_idx is not None and not auto_detect:
                hi = int(header_row_idx)
            elif auto_detect:
                hi = _detect_header_row(rows, max_search=15, min_keywords=2)
            else:
                hi = 0

            header_row = rows[hi]
            ncols = _column_count(header_row)
            headers = []
            for c in range(ncols):
                v = _cell_value(header_row, c)
                key = _normalize_header_cell(v) or f"Column_{c+1}"
                # Dedupe duplicate headers
                base = key
                cnt = 0
                while key in headers:
                    cnt += 1
                    key = f"{base}_{cnt}"
                headers.append(key)

            for row in rows[hi + 1 :]:
                vals = [_cell_value(row, c) for c in range(ncols)]
                if not any(v != "" and v is not None for v in vals):
                    continue
                yield dict(zip(headers, vals))
        finally:
            wb.close()

    def _get_sheet(self, wb: Any, options: dict[str, Any]) -> Any:
        sheet_ref = options.get("sheet")
        if sheet_ref is None:
            return wb.active
        if isinstance(sheet_ref, int):
            return wb.worksheets[sheet_ref]
        return wb[sheet_ref]

    def probe(self, source_path: Path, options: dict[str, Any]) -> SourceProbe:
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("XLSX support requires openpyxl. Install with: pip install openpyxl") from e

        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet = self._get_sheet(wb, options)
            skip_rows = int(options.get("skip_rows", 0))
            header_row_idx = options.get("header_row")
            auto_detect = options.get("auto_detect_header", True)

            rows = list(sheet.iter_rows(min_row=1 + skip_rows, max_row=500))
            if not rows:
                return SourceProbe(row_count=0, columns=(), sample_rows=(), encoding=None, detected_delimiter=None)

            if header_row_idx is not None and not auto_detect:
                hi = int(header_row_idx)
            elif auto_detect:
                hi = _detect_header_row(rows, max_search=15, min_keywords=2)
            else:
                hi = 0

            header_row = rows[hi]
            ncols = _column_count(header_row)
            headers = []
            for c in range(ncols):
                v = _cell_value(header_row, c)
                key = _normalize_header_cell(v) or f"Column_{c+1}"
                base = key
                cnt = 0
                while key in headers:
                    cnt += 1
                    key = f"{base}_{cnt}"
                headers.append(key)

            columns_tuple = tuple(headers)
            sample = []
            for row in rows[hi + 1 : hi + 6]:
                vals = [_cell_value(row, c) for c in range(ncols)]
                if not any(v != "" and v is not None for v in vals):
                    continue
                sample.append(dict(zip(headers, vals)))
            count = len(rows) - hi - 1
            return SourceProbe(
                row_count=count,
                columns=columns_tuple,
                sample_rows=tuple(sample),
                encoding=None,
                detected_delimiter=None,
            )
        finally:
            wb.close()
