"""
Detect QuickBooks Online export type from filename and (optionally) sheet headers.

QBO export filenames often include the report name (e.g. "Company_Account List.xlsx",
"Journal.xlsx", "General Ledger.xlsx"). We match on lowercase filename patterns first,
then fall back to header-based detection for XLSX.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

# Report type identifier -> used for output JSON key and reader selection
QBO_ACCOUNTS = "accounts"
QBO_CUSTOMERS = "customers"
QBO_VENDORS = "vendors"
QBO_JOURNAL = "journal"
QBO_GL = "general_ledger"
QBO_TRIAL_BALANCE = "trial_balance"
QBO_UNKNOWN = "unknown"

# (pattern, type) - pattern is regex on lowercase filename
_FILENAME_PATTERNS = [
    (r"account\s*list", QBO_ACCOUNTS),
    (r"chart\s*of\s*accounts", QBO_ACCOUNTS),
    (r"customer\s*contact\s*list", QBO_CUSTOMERS),
    (r"customer\s*list", QBO_CUSTOMERS),
    (r"vendor\s*contact\s*list", QBO_VENDORS),
    (r"vendor\s*list", QBO_VENDORS),
    (r"journal", QBO_JOURNAL),
    (r"transaction\s*list", QBO_JOURNAL),
    (r"general\s*ledger", QBO_GL),
    (r"gl\b", QBO_GL),
    (r"trial\s*balance", QBO_TRIAL_BALANCE),
    (r"balance\s*sheet", "balance_sheet"),
    (r"profit\s*and\s*loss", "profit_loss"),
    (r"p\s*&\s*l", "profit_loss"),
]


def detect_qbo_type(path: Path, peek_headers: bool = True) -> str:
    """
    Detect QBO export type for the given file.

    Uses filename patterns first. For XLSX, optionally peeks at the first sheet
    headers to disambiguate (e.g. Journal vs GL both might have Date/Account/Debit/Credit).
    """
    stem = path.stem.lower()
    for pattern, report_type in _FILENAME_PATTERNS:
        if re.search(pattern, stem):
            if report_type in (QBO_JOURNAL, QBO_GL) and peek_headers and path.suffix.lower() == ".xlsx":
                # Optionally refine by headers (only .xlsx; .xls is binary and openpyxl can't read it)
                detected = _detect_journal_vs_gl(path)
                if detected:
                    return detected
            return report_type
    if peek_headers and path.suffix.lower() == ".xlsx":
        return _detect_by_headers(path)
    return QBO_UNKNOWN


def _detect_journal_vs_gl(path: Path) -> str | None:
    """Use first sheet headers to distinguish Journal vs General Ledger."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except BadZipFile:
        # File has .xlsx extension but isn't a valid XLSX (e.g. CSV saved as .xlsx, or corrupted)
        return None
    except Exception:
        return None
    try:
        sheet = wb.active
        rows = list(sheet.iter_rows(min_row=1, max_row=10))
        if not rows:
            return None
        # First row that looks like headers
        for row in rows:
            cells = [str(c.value or "").strip().lower() for c in row[:15]]
            if "account" in cells and ("debit" in cells or "credit" in cells):
                if "balance" in cells and cells[0] in ("account", "account name", "name"):
                    return QBO_GL
                return QBO_JOURNAL
    finally:
        wb.close()
    return None


def _detect_by_headers(path: Path) -> str:
    """Infer type from first sheet header row keywords."""
    try:
        import openpyxl
    except ImportError:
        return QBO_UNKNOWN
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except BadZipFile:
        return QBO_UNKNOWN
    except Exception:
        return QBO_UNKNOWN
    try:
        sheet = wb.active
        row0 = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not row0:
            return QBO_UNKNOWN
        headers = [str(c.value or "").strip().lower() for c in row0[:20]]
        if "full name" in headers and "type" in headers and "detail type" in headers:
            return QBO_ACCOUNTS
        if "vendor" in headers and ("phone" in headers or "email" in headers):
            return QBO_VENDORS
        if "name" in headers and ("email" in headers or "company" in headers) and "vendor" not in headers:
            return QBO_CUSTOMERS
        if "debit" in headers and "credit" in headers and "account" in headers:
            return QBO_JOURNAL if "num" in headers or "transaction" in headers else QBO_GL
    finally:
        wb.close()
    return QBO_UNKNOWN
